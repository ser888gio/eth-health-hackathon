import csv
import json
import logging
import os
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any

import requests

logger = logging.getLogger(__name__)


@dataclass
class SophiaPatientData:
    patient_id: str
    sample_id: str | None
    patient_record: dict | None
    qc_text: str | None
    gene_text: str | None
    coverage_text: str | None
    variants_text: str | None
    sources: dict[str, str]


def _require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise ValueError(f"Missing required environment variable: {name}")
    return value


def _format_url(template: str, base_url: str | None, **kwargs: Any) -> str:
    rendered = template.format(**kwargs)
    if rendered.startswith("http://") or rendered.startswith("https://"):
        return rendered
    if not base_url:
        raise ValueError("SOPHIA_API_BASE_URL must be set when using relative URL templates")
    return f"{base_url.rstrip('/')}/{rendered.lstrip('/')}"


def _ensure_sample_id(template: str, sample_id: str | None, name: str) -> None:
    if "{sample_id}" in template and not sample_id:
        raise ValueError(f"{name} requires sample_id, but none was provided")


def _headers() -> dict[str, str]:
    token = os.getenv("SOPHIA_API_TOKEN", "").strip()
    headers: dict[str, str] = {"Accept": "application/json, text/plain, */*"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def _request_text(url: str) -> str:
    resp = requests.get(url, headers=_headers(), timeout=60)
    resp.raise_for_status()
    return _response_to_text(resp)


def _response_to_text(resp: requests.Response) -> str:
    content_type = (resp.headers.get("Content-Type") or "").lower()
    if "application/json" in content_type:
        return json.dumps(resp.json(), indent=2)

    data = resp.content
    if "spreadsheetml.sheet" in content_type or data.startswith(b"PK"):
        return _xlsx_to_csv_text(data)

    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("latin-1", errors="replace")


def _xlsx_to_csv_text(data: bytes, max_rows: int | None = None) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to parse .xlsx exports. "
            "Install it or export CSV from SOPHiA instead."
        ) from exc

    max_rows_env = os.getenv("SOPHIA_MAX_ROWS", "")
    if max_rows_env.strip().isdigit():
        max_rows = int(max_rows_env)

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    output = StringIO()
    writer = csv.writer(output)

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows is not None and idx >= max_rows:
            writer.writerow(["...TRUNCATED..."])
            break
        writer.writerow(["" if cell is None else cell for cell in row])

    return output.getvalue()


def _truncate_text(text: str) -> str:
    max_chars = int(os.getenv("SOPHIA_MAX_TEXT_CHARS", "200000"))
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...TRUNCATED..."


def fetch_patient_bundle(patient_id: str, sample_id: str | None = None) -> SophiaPatientData:
    base_url = os.getenv("SOPHIA_API_BASE_URL", "").strip() or None

    patient_template = _require_env("SOPHIA_PATIENT_URL_TEMPLATE")
    variants_template = _require_env("SOPHIA_VARIANTS_URL_TEMPLATE")
    coverage_template = _require_env("SOPHIA_COVERAGE_URL_TEMPLATE")

    _ensure_sample_id(patient_template, sample_id, "SOPHIA_PATIENT_URL_TEMPLATE")
    _ensure_sample_id(variants_template, sample_id, "SOPHIA_VARIANTS_URL_TEMPLATE")
    _ensure_sample_id(coverage_template, sample_id, "SOPHIA_COVERAGE_URL_TEMPLATE")

    patient_url = _format_url(patient_template, base_url, patient_id=patient_id, sample_id=sample_id or "")
    variants_url = _format_url(variants_template, base_url, patient_id=patient_id, sample_id=sample_id or "")
    coverage_url = _format_url(coverage_template, base_url, patient_id=patient_id, sample_id=sample_id or "")

    logger.info("Fetching patient metadata: %s", patient_url)
    patient_raw = _request_text(patient_url)
    patient_record: dict | None
    try:
        patient_record = json.loads(patient_raw)
    except json.JSONDecodeError:
        patient_record = {"raw": patient_raw}

    logger.info("Fetching variant export: %s", variants_url)
    variants_text = _truncate_text(_request_text(variants_url))

    logger.info("Fetching coverage export: %s", coverage_url)
    coverage_text = _truncate_text(_request_text(coverage_url))

    return SophiaPatientData(
        patient_id=patient_id,
        sample_id=sample_id,
        patient_record=patient_record,
        qc_text=None,
        gene_text=None,
        coverage_text=coverage_text,
        variants_text=variants_text,
        sources={
            "patient_url": patient_url,
            "variants_url": variants_url,
            "coverage_url": coverage_url,
        },
    )
